"""
Script que convierte el Excel del RENAJU → orgs_cusco.json
Filtro: Region == Cusco
Coordenadas: mapeo por provincia (centroide conocido)
"""
import json
import os
import re
import openpyxl

EXCEL = os.path.join(os.path.dirname(__file__), "..", "BASE-DE-DATOS-ORGANIZACIONES-JUVENILES-E-INSTITUCIONES-PRIVADAS.xlsx")
OUTPUT = os.path.join(os.path.dirname(__file__), "data", "orgs_cusco.json")

# Coordenadas aproximadas por provincia de Cusco
COORDS = {
    "Cusco":         (-13.5170, -71.9785),
    "Acomayo":       (-13.9203, -71.6917),
    "Anta":          (-13.4757, -72.1494),
    "Calca":         (-13.3322, -71.9636),
    "Canas":         (-14.4000, -71.4500),
    "Canchis":       (-14.1667, -71.2000),
    "Chumbivilcas":  (-14.4333, -72.0833),
    "Espinar":       (-14.7897, -71.4061),
    "La Convención": (-12.8500, -72.7000),
    "Paruro":        (-13.7517, -71.8417),
    "Paucartambo":   (-13.3167, -71.5917),
    "Quispicanchi":  (-13.6167, -71.5333),
    "Urubamba":      (-13.3003, -72.1170),
}

# Acción concreta según temática
def accion(tematica1: str, tipo: str) -> str:
    t = (tematica1 or "").lower()
    if "medio" in t or "naturaleza" in t or "recursos" in t:
        return "Únete a la próxima jornada de limpieza o reforestación. Escríbeles por email para conocer su calendario de actividades."
    if "educaci" in t:
        return "Participa como tutor voluntario o asiste a sus talleres comunitarios. Contáctalos por email esta semana."
    if "salud" in t:
        return "Colabora en sus campañas de salud preventiva. Escríbeles para conocer su próxima actividad."
    if "democracia" in t or "derechos" in t or "paz" in t:
        return "Participa en sus talleres de ciudadanía y derechos. Contáctalos para sumarte como voluntario."
    if "cultura" in t or "arte" in t:
        return "Súmate a sus actividades culturales o artísticas. Escríbeles para conocer el calendario."
    if "ciencia" in t or "tecnolog" in t or "tic" in t:
        return "Contribuye con tus habilidades técnicas en sus proyectos. Contáctalos esta semana."
    if "género" in t or "sexual" in t or "reproduct" in t:
        return "Participa en sus espacios de formación y diálogo. Escríbeles para unirte."
    if "solidar" in t or "altruist" in t or "voluntar" in t:
        return "Regístrate como voluntario y participa en su próxima actividad comunitaria."
    if "desarrollo" in t or "económ" in t:
        return "Participa en sus talleres de desarrollo comunitario. Contáctalos para más información."
    if "investigaci" in t:
        return "Colabora en sus proyectos de investigación y documentación. Escríbeles por email."
    if "comunicaci" in t:
        return "Únete como colaborador en sus medios y campañas comunitarias. Contáctalos."
    return "Escríbeles por email para conocer cómo participar esta semana en sus actividades."


def main():
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    orgs = []
    idx = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        region = str(d.get("Region") or "").strip()
        if "cusco" not in region.lower():
            continue

        provincia = str(d.get("Provincia") or "").strip().rstrip()
        # Normalizar "Cusco " → "Cusco"
        provincia_norm = re.sub(r"\s+", " ", provincia).strip()

        lat, lng = COORDS.get(provincia_norm, COORDS["Cusco"])

        orgs.append({
            "id": idx,
            "nombre": str(d.get("Organización Juvenil") or "").strip(),
            "region": "Cusco",
            "provincia": provincia_norm,
            "representante": str(d.get("Representante") or "").strip(),
            "email": str(d.get("Email de la organización") or "").strip(),
            "tipo": str(d.get("Tipo de organización ") or "").strip(),
            "tematica1": str(d.get("Temática de organización 1") or "").strip(),
            "tematica2": str(d.get("Temática de organización 2") or "").strip(),
            "descripcion": f"Organización juvenil de {provincia_norm}, Cusco, enfocada en {str(d.get('Temática de organización 1') or '').lower()}.",
            "lat": lat,
            "lng": lng,
            "accion_concreta": accion(
                str(d.get("Temática de organización 1") or ""),
                str(d.get("Tipo de organización ") or "")
            ),
        })
        idx += 1

    with open(OUTPUT, "w", encoding="utf-8") as f:
        json.dump(orgs, f, ensure_ascii=False, indent=2)

    print(f"Generadas {len(orgs)} orgs → {OUTPUT}")


if __name__ == "__main__":
    main()
